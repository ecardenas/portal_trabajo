import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the base data
df = pd.read_csv('Caso de Uso.csv')
num_rows = len(df)

# Create Workbook
wb = Workbook()
ws_dash = wb.active
ws_dash.title = "Dashboard Ejecutivo"
ws_data = wb.create_sheet("Matriz de Datos y Controles")

# Colors & Styles
bg_color = "EDEEED"
header_color = "1F4E78" # Dark Blue
white_text = "FFFFFF"
border_style = Side(style='thin', color="000000")

# --- STYLE DASHBOARD ---
for row in range(1, 40):
    for col in range(1, 15):
        ws_dash.cell(row=row, column=col).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

# Dashboard Title
ws_dash.merge_cells('B2:M3')
title_cell = ws_dash['B2']
title_cell.value = "REPORTE EJECUTIVO DE CALIDAD DE DATOS - COMER PERÚ"
title_cell.font = Font(name='Arial', size=18, bold=True, color=white_text)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

# KPI Setup
kpis = [
    ("COMPLETITUD", "B5:D7", "='Matriz de Datos y Controles'!$Q$2"),
    ("EXACTITUD", "E5:G7", "='Matriz de Datos y Controles'!$R$2"),
    ("UNICIDAD", "H5:J7", "='Matriz de Datos y Controles'!$S$2"),
    ("VALIDEZ", "K5:M7", "='Matriz de Datos y Controles'!$T$2")
]

for label, merge_range, formula in kpis:
    ws_dash.merge_cells(merge_range)
    cell = ws_dash[merge_range.split(':')[0]]
    cell.value = f"{label}\n{formula}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.number_format = '0.0%'

# Global Index
ws_dash.merge_cells('B10:M11')
idx_cell = ws_dash['B10']
idx_cell.value = "ÍNDICE DE CALIDAD GLOBAL (ICG)\n=(B6+E6+H6+K6)/4"
idx_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
idx_cell.number_format = '0.0%'
idx_cell.font = Font(size=24, bold=True)

# Veredicto
ws_dash.merge_cells('B13:M13')
verdict_cell = ws_dash['B13']
verdict_cell.value = '=IF(B11>0.85; "ESTADO: ÓPTIMO"; IF(B11>0.7; "ESTADO: ACEPTABLE CON OBSERVACIONES"; "ESTADO: CRÍTICO - REQUIERE LIMPIEZA"))'
verdict_cell.font = Font(bold=True, size=14)
verdict_cell.alignment = Alignment(horizontal='center')

# --- DATA SHEET ---
headers = list(df.columns) + ["CONT: Completitud", "CONT: Exactitud", "CONT: Unicidad", "CONT: Validez"]
for col_num, header in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color=white_text)
    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

# Ajustar ancho de columnas automáticamente
for i, col in enumerate(ws_data.columns, 1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_data.column_dimensions[get_column_letter(i)].width = max_length + 2

# Guardar el archivo
wb.save("reporte_calidad.xlsx")

# Summary formulas at the top of Data Sheet (Row 2)
ws_data['Q1'] = "Promedio Completitud"
ws_data['Q2'] = f"=AVERAGE(Q21:Q{20+num_rows})"
ws_data['R1'] = "Promedio Exactitud"
ws_data['R2'] = f"=AVERAGE(R21:R{20+num_rows})"
ws_data['S1'] = "Promedio Unicidad"
ws_data['S2'] = f"=AVERAGE(S21:S{20+num_rows})"
ws_data['T1'] = "Promedio Validez"
ws_data['T2'] = f"=AVERAGE(T21:T{20+num_rows})"

# Populate Data
for r_idx, row in enumerate(df.values, 21):
    for c_idx, value in enumerate(row, 1):
        ws_data.cell(row=r_idx, column=c_idx).value = value
    
    # Dynamic Rules Formulas (using Row relative refs)
    # Completitud: Check ID and RUC not empty/placeholders
    ws_data.cell(row=r_idx, column=17).value = f'=IF(AND(A{r_idx}<>""; B{r_idx}<>""; C{r_idx}<>"###"); 1; 0)'
    
    # Exactitud: Monto > 0 and numeric
    ws_data.cell(row=r_idx, column=18).value = f'=IF(AND(ISNUMBER(D{r_idx}); D{r_idx}>0); 1; 0)'
    
    # Unicidad: ID Client unique in range
    ws_data.cell(row=r_idx, column=19).value = f'=IF(COUNTIF($A$21:$A${20+num_rows}; A{r_idx})=1; 1; 0)'
    
    # Validez: RUC length 11 and numeric-ish (start with 10 or 20)
    ws_data.cell(row=r_idx, column=20).value = f'=IF(AND(LEN(B{r_idx})=11; OR(LEFT(B{r_idx};2)="10"; LEFT(B{r_idx};2)="20")); 1; 0)'

# Save file
output_file = "Auditoria_Calidad_ComerPeru_v3.xlsx"
wb.save(output_file)